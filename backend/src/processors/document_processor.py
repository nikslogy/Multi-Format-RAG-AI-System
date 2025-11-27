"""
Document Processor Module
Handles extraction and chunking of text from multiple file formats.
Includes table extraction and contextual chunk headers.
"""

import os
import json
from datetime import datetime
from typing import List, Dict, Tuple
import PyPDF2
from docx import Document
import openpyxl
import pandas as pd

# Table extraction libraries (optional, graceful degradation)
try:
    import camelot
    CAMELOT_AVAILABLE = True
except ImportError:
    CAMELOT_AVAILABLE = False
    print("Warning: camelot-py not installed. PDF table extraction will be limited.")

try:
    import tabula
    TABULA_AVAILABLE = True
except ImportError:
    TABULA_AVAILABLE = False
    print("Warning: tabula-py not installed. PDF table extraction will use fallback.")


class DocumentProcessor:
    """Process documents of various formats and extract text content."""

    def __init__(self, chunk_size: int = 500, chunk_overlap: int = 100):
        """
        Initialize the document processor.

        Args:
            chunk_size: Number of words per chunk
            chunk_overlap: Number of words to overlap between chunks
        """
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    def extract_text(self, file_path: str, filename: str) -> str:
        """
        Extract text from a file based on its extension.

        Args:
            file_path: Path to the file
            filename: Original filename with extension

        Returns:
            Extracted text content
        """
        extension = os.path.splitext(filename)[1].lower()

        try:
            if extension == '.pdf':
                return self._extract_from_pdf(file_path)
            elif extension == '.docx':
                return self._extract_from_docx(file_path)
            elif extension == '.txt':
                return self._extract_from_txt(file_path)
            elif extension == '.xlsx':
                return self._extract_from_xlsx(file_path)
            elif extension == '.json':
                return self._extract_from_json(file_path)
            elif extension == '.csv':
                return self._extract_from_csv(file_path)
            else:
                raise ValueError(f"Unsupported file format: {extension}")
        except Exception as e:
            raise Exception(f"Error extracting text from {filename}: {str(e)}")

    def _extract_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file."""
        text = []
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                text.append(page.extract_text())
        return '\n'.join(text)

    def _extract_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file."""
        doc = Document(file_path)
        return '\n'.join([paragraph.text for paragraph in doc.paragraphs])

    def _extract_from_txt(self, file_path: str) -> str:
        """Extract text from TXT file."""
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            return file.read()

    def _extract_from_xlsx(self, file_path: str) -> str:
        """Extract text from XLSX file."""
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        text = []

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text.append(f"Sheet: {sheet_name}")

                for row in sheet.iter_rows(values_only=True):
                    row_text = ' | '.join([str(cell) if cell is not None else '' for cell in row])
                    if row_text.strip():
                        text.append(row_text)

            return '\n'.join(text)
        finally:
            # Ensure workbook is closed to release file handle (important on Windows)
            workbook.close()

    def _extract_from_json(self, file_path: str) -> str:
        """Extract text from JSON file."""
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return json.dumps(data, indent=2)

    def _extract_from_csv(self, file_path: str) -> str:
        """Extract text from CSV file."""
        df = pd.read_csv(file_path)
        return df.to_string(index=False)

    def extract_tables_from_pdf(self, file_path: str, filename: str) -> List[Dict[str, any]]:
        """
        Extract tables from PDF files.

        Args:
            file_path: Path to PDF file
            filename: Original filename

        Returns:
            List of table chunks with metadata
        """
        tables = []

        try:
            if CAMELOT_AVAILABLE:
                # Use Camelot for better table extraction
                extracted_tables = camelot.read_pdf(file_path, pages='all', flavor='lattice')

                for i, table in enumerate(extracted_tables):
                    df = table.df

                    # Create text representation with summary statistics
                    text_representation = f"Table {i+1} from {filename}:\n{df.to_string(index=False)}"

                    # Add numerical summaries if table contains numbers
                    numeric_cols = df.select_dtypes(include=['number']).columns
                    if len(numeric_cols) > 0:
                        summary = f"\n\nNumerical Summary:\n{df[numeric_cols].describe().to_string()}"
                        text_representation += summary

                    # Create table chunk
                    table_chunk = {
                        'text': f"Document: {filename} | Type: TABLE {i+1} | Content: {text_representation}",
                        'original_text': text_representation,
                        'metadata': {
                            'filename': filename,
                            'chunk_number': i,
                            'upload_date': datetime.now().isoformat(),
                            'word_count': len(text_representation.split()),
                            'content_type': 'table',
                            'table_index': i,
                            'table_rows': df.shape[0],  # Number of rows
                            'table_cols': df.shape[1],  # Number of columns
                            'has_numerical_data': len(numeric_cols) > 0
                        },
                        'structured_data': df.to_dict()  # Store structured data for calculations
                    }
                    tables.append(table_chunk)

            elif TABULA_AVAILABLE:
                # Fallback to Tabula
                extracted_tables = tabula.read_pdf(file_path, pages='all', multiple_tables=True)

                for i, df in enumerate(extracted_tables):
                    if df.empty:
                        continue

                    text_representation = f"Table {i+1} from {filename}:\n{df.to_string(index=False)}"

                    numeric_cols = df.select_dtypes(include=['number']).columns
                    if len(numeric_cols) > 0:
                        summary = f"\n\nNumerical Summary:\n{df[numeric_cols].describe().to_string()}"
                        text_representation += summary

                    table_chunk = {
                        'text': f"Document: {filename} | Type: TABLE {i+1} | Content: {text_representation}",
                        'original_text': text_representation,
                        'metadata': {
                            'filename': filename,
                            'chunk_number': i,
                            'upload_date': datetime.now().isoformat(),
                            'word_count': len(text_representation.split()),
                            'content_type': 'table',
                            'table_index': i,
                            'table_rows': df.shape[0],  # Number of rows
                            'table_cols': df.shape[1],  # Number of columns
                            'has_numerical_data': len(numeric_cols) > 0
                        },
                        'structured_data': df.to_dict()
                    }
                    tables.append(table_chunk)

        except Exception as e:
            print(f"Warning: Could not extract tables from {filename}: {str(e)}")

        return tables

    def extract_tables_from_excel(self, file_path: str, filename: str) -> List[Dict[str, any]]:
        """
        Extract tables/sheets from Excel files with numerical summaries.

        Args:
            file_path: Path to Excel file
            filename: Original filename

        Returns:
            List of table chunks with metadata
        """
        tables = []
        excel_file = None

        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)

            for sheet_idx, sheet_name in enumerate(excel_file.sheet_names):
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                if df.empty:
                    continue

                # Create text representation
                text_representation = f"Sheet '{sheet_name}' from {filename}:\n{df.to_string(index=False)}"

                # Add numerical summaries
                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    summary = f"\n\nNumerical Summary for {sheet_name}:\n{df[numeric_cols].describe().to_string()}"
                    text_representation += summary

                    # Add totals for each numeric column
                    totals = df[numeric_cols].sum()
                    text_representation += f"\n\nColumn Totals:\n{totals.to_string()}"

                # Create table chunk
                table_chunk = {
                    'text': f"Document: {filename} | Type: TABLE (Sheet: {sheet_name}) | Content: {text_representation}",
                    'original_text': text_representation,
                    'metadata': {
                        'filename': filename,
                        'chunk_number': sheet_idx,
                        'upload_date': datetime.now().isoformat(),
                        'word_count': len(text_representation.split()),
                        'content_type': 'table',
                        'sheet_name': sheet_name,
                        'table_index': sheet_idx,
                        'table_rows': df.shape[0],  # Number of rows
                        'table_cols': df.shape[1],  # Number of columns
                        'has_numerical_data': len(numeric_cols) > 0
                    },
                    'structured_data': df.to_dict()  # For potential calculations
                }
                tables.append(table_chunk)

        except Exception as e:
            print(f"Warning: Could not extract tables from {filename}: {str(e)}")

        finally:
            # Ensure Excel file is closed to release file handle (important on Windows)
            if excel_file is not None:
                excel_file.close()

        return tables

    def chunk_text(self, text: str, filename: str, add_context_header: bool = True) -> List[Dict[str, any]]:
        """
        Split text into overlapping chunks with metadata and contextual headers.

        Args:
            text: Text to chunk
            filename: Source filename
            add_context_header: Whether to add contextual header to each chunk

        Returns:
            List of chunks with metadata
        """
        if not text or not text.strip():
            return []

        # Split into words
        words = text.split()
        chunks = []

        # Create overlapping chunks
        start_idx = 0
        chunk_num = 0

        while start_idx < len(words):
            # Get chunk of words
            end_idx = start_idx + self.chunk_size
            chunk_words = words[start_idx:end_idx]
            chunk_text = ' '.join(chunk_words)

            # Add contextual header for better retrieval
            if add_context_header:
                # Extract document name without extension
                doc_name = os.path.splitext(filename)[0]

                # Create context header
                context_header = f"Document: {filename} | Section: Part {chunk_num + 1} | Content: "
                chunk_text_with_context = context_header + chunk_text
            else:
                chunk_text_with_context = chunk_text

            # Create chunk with metadata
            chunk = {
                'text': chunk_text_with_context,  # Text with contextual header
                'original_text': chunk_text,  # Original text without header
                'metadata': {
                    'filename': filename,
                    'chunk_number': chunk_num,
                    'upload_date': datetime.now().isoformat(),
                    'word_count': len(chunk_words),
                    'start_word': start_idx,
                    'end_word': min(end_idx, len(words)),
                    'content_type': 'text'  # To distinguish from tables
                }
            }
            chunks.append(chunk)

            # Move to next chunk with overlap
            start_idx += (self.chunk_size - self.chunk_overlap)
            chunk_num += 1

        return chunks

    def process_document(self, file_path: str, filename: str) -> Tuple[List[str], List[dict]]:
        """
        Complete pipeline: extract text, tables, and create chunks.

        Args:
            file_path: Path to the file
            filename: Original filename

        Returns:
            Tuple of (chunk_texts, metadatas) - lists of strings and metadata dicts
        """
        extension = os.path.splitext(filename)[1].lower()

        # Extract text from document
        text = self.extract_text(file_path, filename)

        # Validate text extraction
        if not text or len(text.strip()) < 10:
            raise ValueError(f"No meaningful text extracted from {filename}")

        # Create text chunks with contextual headers
        text_chunks = self.chunk_text(text, filename)

        # Extract tables based on file type
        table_chunks = []
        if extension == '.pdf':
            table_chunks = self.extract_tables_from_pdf(file_path, filename)
        elif extension in ['.xlsx', '.xls']:
            table_chunks = self.extract_tables_from_excel(file_path, filename)
        elif extension == '.csv':
            # CSV is already treated as a table, create enhanced version
            try:
                df = pd.read_csv(file_path)
                text_rep = f"CSV Data from {filename}:\n{df.to_string(index=False)}"

                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    summary = f"\n\nNumerical Summary:\n{df[numeric_cols].describe().to_string()}"
                    totals = df[numeric_cols].sum()
                    text_rep += summary + f"\n\nColumn Totals:\n{totals.to_string()}"

                table_chunk = {
                    'text': f"Document: {filename} | Type: CSV TABLE | Content: {text_rep}",
                    'original_text': text_rep,
                    'metadata': {
                        'filename': filename,
                        'chunk_number': 0,
                        'upload_date': datetime.now().isoformat(),
                        'word_count': len(text_rep.split()),
                        'content_type': 'table',
                        'table_index': 0,
                        'table_rows': df.shape[0],
                        'table_cols': df.shape[1],
                        'has_numerical_data': len(numeric_cols) > 0
                    },
                    'structured_data': df.to_dict()
                }
                table_chunks.append(table_chunk)
            except Exception as e:
                print(f"Warning: Could not process CSV as table: {str(e)}")

        # Convert to format expected by RAG engine (separate lists)
        all_chunks = text_chunks + table_chunks
        chunk_texts = [chunk['text'] for chunk in all_chunks]
        metadatas = [chunk['metadata'] for chunk in all_chunks]

        return chunk_texts, metadatas

    def create_excel_metadata(self, file_path: str, filename: str) -> Tuple[List[str], List[dict]]:
        """
        Create a metadata-only chunk for Excel/CSV files.
        Instead of chunking all data, stores file info for code execution.

        Args:
            file_path: Path to the Excel/CSV file
            filename: Original filename

        Returns:
            Tuple of (chunks, metadatas) for consistency with process_document
        """
        extension = os.path.splitext(filename)[1].lower()

        try:
            if extension == '.csv':
                df = pd.read_csv(file_path)
                sheet_info = "CSV file (single table)"
            else:
                # For Excel files, read first sheet
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                df = pd.read_excel(excel_file, sheet_name=sheet_names[0])
                sheet_info = f"Excel file with {len(sheet_names)} sheet(s): {', '.join(sheet_names)}"
                excel_file.close()

            # Get column information
            columns = df.columns.tolist()
            dtypes = df.dtypes.to_dict()

            # Get sample data (first 10 rows)
            sample_data = df.head(10).to_string(index=False)

            # Get basic statistics for numeric columns
            numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
            stats_info = ""
            if numeric_cols:
                stats_info = f"\n\nNumeric columns: {', '.join(numeric_cols)}"

            # Create informative metadata text
            metadata_text = f"""
Excel/CSV File: {filename}

Structure:
- {sheet_info}
- Total Rows: {len(df)}
- Total Columns: {len(columns)}
- Columns: {', '.join(columns)}
{stats_info}

Sample Data (first 10 rows):
{sample_data}

NOTE: This is a structured data file stored for code execution analysis.
For queries about calculations, aggregations, filtering, or data analysis,
the system will use Python code execution to analyze the full dataset.
            """.strip()

            # Create metadata chunk
            chunk_text = f"Document: {filename} | Type: EXCEL METADATA | Content: {metadata_text}"

            chunk_metadata = {
                'filename': filename,
                'chunk_number': 0,
                'upload_date': datetime.now().isoformat(),
                'word_count': len(metadata_text.split()),
                'content_type': 'excel_metadata',
                'file_path': file_path,
                'columns': ', '.join(columns),
                'row_count': len(df),
                'col_count': len(columns),
                'has_numerical_data': len(numeric_cols) > 0,
                'numeric_columns': ', '.join(numeric_cols) if numeric_cols else ''
            }

            # Return as tuple (chunks, metadatas) for consistency
            return ([chunk_text], [chunk_metadata])

        except Exception as e:
            raise Exception(f"Error creating Excel metadata for {filename}: {str(e)}")
